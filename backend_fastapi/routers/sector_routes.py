"""
Sector Routes
=============

Handles extraction of consumption sectors from Excel files.

Endpoints:
- GET /project/sectors - Extract sector names from input_demand_file.xlsx
"""

from fastapi import APIRouter, HTTPException, Query
from pathlib import Path
import openpyxl
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/sectors")
async def get_sectors(projectPath: str = Query(..., description="Project root path")):
    """
    Extract consumption sector names from the main Excel file.

    Reads the 'main' sheet and looks for the '~consumption_sectors' marker,
    then extracts all sector names listed below it.

    Args:
        projectPath: Project root directory

    Returns:
        dict: List of sector names

    Raises:
        HTTPException: 400 if path missing, 404 if file not found, 500 on error
    """
    if not projectPath:
        raise HTTPException(
            status_code=400,
            detail="Project path is missing"
        )

    try:
        file_path = Path(projectPath) / "inputs" / "input_demand_file.xlsx"

        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found at path: {file_path}"
            )

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = workbook.sheetnames[0]
        worksheet = workbook[sheet_name]

        # Convert to list of rows
        rows = list(worksheet.iter_rows(values_only=True))

        sectors = []
        start_index = -1

        # Find the marker '~consumption_sectors' (case-insensitive)
        for i, row in enumerate(rows):
            for cell_value in row:
                if (isinstance(cell_value, str) and
                    cell_value.strip().lower() == '~consumption_sectors'):
                    start_index = i + 2  # Start reading 2 rows below the marker
                    break
            if start_index != -1:
                break

        # Extract sectors starting from the marker
        if start_index != -1:
            for i in range(start_index, len(rows)):
                cell = rows[i][0]  # First column
                if not cell or str(cell).strip() == '':
                    continue
                sectors.append(str(cell).strip())

        workbook.close()

        return {"sectors": sectors}

    except HTTPException:
        raise
    except Exception as err:
        logger.error(f"❌ Error reading Excel file: {err}")
        raise HTTPException(
            status_code=500,
            detail="Failed to read Excel file"
        )
