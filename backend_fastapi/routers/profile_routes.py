"""
Load Profile Generation Routes
==============================

Handles load profile generation with real-time progress via SSE.

Endpoints:
- GET /project/available-base-years - List financial years from load curve template
- GET /project/available-scenarios - List completed demand forecast scenarios
- POST /project/generate-profile - Start profile generation process
- GET /project/generation-status - Server-Sent Events for generation progress
- GET /project/check-profile-exists - Check if a profile file already exists
"""

import subprocess
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
import openpyxl
import asyncio
import json
import logging
import threading
import queue

logger = logging.getLogger(__name__)
router = APIRouter()

# Global queue for SSE events
profile_event_queue: asyncio.Queue = None


def get_financial_year(date: datetime) -> str:
    """
    Calculate financial year from a date (Apr-Mar cycle).

    Args:
        date: Date object

    Returns:
        Financial year string (e.g., 'FY2024')
    """
    year = date.year
    month = date.month
    financial_year = year + 1 if month >= 4 else year  # April is month 4
    return f"FY{financial_year}"


class ProfileConfiguration(BaseModel):
    """Profile generation configuration"""
    # Add fields as needed based on the actual configuration structure
    pass


class GenerateProfileRequest(BaseModel):
    """Request model for profile generation"""
    projectPath: str = Field(..., description="Project root path")
    profileConfiguration: Dict[str, Any] = Field(..., description="Profile configuration data")


@router.get("/available-base-years")
async def get_available_base_years(projectPath: str = Query(..., description="Project root path")):
    """
    Extract unique financial years from the load curve template.

    Reads the 'Past_Hourly_Demand' sheet and extracts FY values from dates.

    Args:
        projectPath: Project root directory

    Returns:
        dict: List of financial year strings (e.g., ['FY2021', 'FY2022'])
    """
    if not projectPath:
        raise HTTPException(status_code=400, detail="Project path is required.")

    try:
        file_path = Path(projectPath) / "inputs" / "load_curve_template.xlsx"

        if not file_path.exists():
            raise HTTPException(status_code=404, detail="load_curve_template.xlsx not found.")

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Find sheet (case-insensitive)
        sheet_name = next(
            (name for name in workbook.sheetnames if name.lower() == 'past_hourly_demand'),
            None
        )

        if not sheet_name:
            workbook.close()
            raise HTTPException(status_code=404, detail="Sheet 'Past_Hourly_Demand' not found.")

        worksheet = workbook[sheet_name]

        # Read data
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        if len(data) == 0:
            workbook.close()
            return {"success": True, "years": []}

        # Find date column (case-insensitive)
        date_header = next(
            (key for key in headers if key and key.lower() == 'date'),
            None
        )

        if not date_header:
            workbook.close()
            raise HTTPException(status_code=404, detail="Column 'date' not found in 'Past_Hourly_Demand' sheet.")

        # Extract financial years
        financial_years = set()
        for row in data:
            date_value = row.get(date_header)
            if date_value:
                # Handle both datetime objects and string dates
                if isinstance(date_value, datetime):
                    parsed_date = date_value
                else:
                    try:
                        parsed_date = datetime.fromisoformat(str(date_value))
                    except:
                        continue

                if parsed_date:
                    financial_years.add(get_financial_year(parsed_date))

        workbook.close()

        # Sort years
        sorted_years = sorted(financial_years, key=lambda x: int(x[2:]))  # Sort by year number

        return {"success": True, "years": sorted_years}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"❌ Error fetching base years: {error}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching base years.")


@router.get("/available-scenarios")
async def get_available_scenarios(projectPath: str = Query(..., description="Project root path")):
    """
    List completed demand forecast scenarios (those with Consolidated_Results.xlsx).

    Args:
        projectPath: Project root directory

    Returns:
        dict: List of scenario names that have consolidated results
    """
    if not projectPath:
        raise HTTPException(status_code=400, detail="Project path is required.")

    try:
        scenarios_parent_path = Path(projectPath) / "results" / "demand_forecasts"

        if not scenarios_parent_path.exists():
            return {"success": True, "scenarios": []}

        # Find directories with Consolidated_Results.xlsx
        valid_scenarios = []
        for item in scenarios_parent_path.iterdir():
            if not item.is_dir():
                continue

            expected_file = item / "Consolidated_Results.xlsx"
            if expected_file.exists():
                valid_scenarios.append(item.name)

        return {"success": True, "scenarios": sorted(valid_scenarios)}

    except Exception as error:
        logger.error(f"❌ Error fetching available scenarios: {error}")
        raise HTTPException(status_code=500, detail="Internal server error while fetching scenarios.")


@router.get("/check-profile-exists")
async def check_profile_exists(
    projectPath: str = Query(..., description="Project root path"),
    profileName: str = Query(..., description="Profile name")
):
    """
    Check if a load profile file already exists.

    Args:
        projectPath: Project root directory
        profileName: Name of the profile (without .xlsx extension)

    Returns:
        dict: Boolean indicating file existence
    """
    if not projectPath or not profileName:
        raise HTTPException(status_code=400, detail="Project path and profile name are required.")

    try:
        file_path = Path(projectPath) / "results" / "load_profiles" / f"{profileName}.xlsx"
        exists = file_path.exists()
        return {"exists": exists}

    except Exception as error:
        logger.error(f"❌ Error checking profile existence: {error}")
        raise HTTPException(status_code=500, detail="Internal server error while checking profile existence.")


@router.post("/generate-profile", status_code=202)
async def generate_profile(request: GenerateProfileRequest):
    """
    Start load profile generation process.

    Spawns a Python subprocess to run load_profile_generation.py with the provided configuration.
    Progress updates are sent via Server-Sent Events to /generation-status endpoint.

    Args:
        request: Profile generation configuration

    Returns:
        dict: Success message (202 Accepted)
    """
    global profile_event_queue

    if not request.projectPath or not request.profileConfiguration:
        raise HTTPException(
            status_code=400,
            detail="Both 'projectPath' and 'profileConfiguration' are required in the request body."
        )

    # Initialize event queue
    profile_event_queue = asyncio.Queue()

    # Prepare full configuration
    full_config = {
        "project_path": request.projectPath,
        "profile_configuration": request.profileConfiguration
    }

    # Start Python process in background
    asyncio.create_task(run_profile_generation_process(full_config, profile_event_queue))

    return {
        "success": True,
        "message": "Generation process started successfully."
    }


async def run_profile_generation_process(config: dict, event_queue: asyncio.Queue):
    """
    Run the Python load profile generation script as a subprocess using threading.

    Args:
        config: Configuration dictionary
        event_queue: Queue for sending SSE events
    """
    python_script_path = Path(__file__).parent.parent / "models" / "load_profile_generation.py"
    config_string = json.dumps(config)
    logger.info(f"Starting profile generation process with script: {python_script_path}")
    logger.info(f"Config: {config_string}")

    def run_subprocess():
        """Run the subprocess in a separate thread"""
        try:
            # Check if script exists
            if not python_script_path.exists():
                asyncio.run(event_queue.put({
                    "type": "error",
                    "message": f"Script not found: {python_script_path}"
                }))
                asyncio.run(event_queue.put({"type": "done"}))
                return

            logger.info("Starting profile generation subprocess...")

            # Start subprocess using synchronous subprocess
            process = subprocess.Popen(
                ["python", str(python_script_path), "--config", config_string],
                cwd=str(python_script_path.parent),  # Set working directory
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,  # Use text mode for easier string handling
                bufsize=1   # Line buffered
            )

            logger.info(f"Profile generation subprocess started with PID: {process.pid}")

            # Read stdout and stderr in separate threads
            final_json_output = ""

            def read_stdout():
                nonlocal final_json_output
                try:
                    for line in iter(process.stdout.readline, ''):
                        line = line.strip()
                        if line:
                            logger.info(f"[Profile Generation STDOUT]: {line}")
                            # Capture final JSON output
                            final_json_output = line
                except Exception as e:
                    logger.error(f"Error reading profile generation stdout: {e}")

            def read_stderr():
                try:
                    for line in iter(process.stderr.readline, ''):
                        line = line.strip()
                        if line:
                            logger.error(f"[Profile Generation STDERR]: {line}")
                            # Send stderr messages as log events
                            asyncio.run(event_queue.put({"type": "log", "data": line}))
                except Exception as e:
                    logger.error(f"Error reading profile generation stderr: {e}")

            # Start reading threads
            stdout_thread = threading.Thread(target=read_stdout, daemon=True)
            stderr_thread = threading.Thread(target=read_stderr, daemon=True)

            stdout_thread.start()
            stderr_thread.start()

            # Wait for process to complete
            process.wait()
            logger.info(f"Profile generation script exited with code {process.returncode}")

            # Wait for reading threads to complete
            stdout_thread.join(timeout=5.0)
            stderr_thread.join(timeout=5.0)

            # Clean up
            process.stdout.close()
            process.stderr.close()

            # Parse and send final result
            if process.returncode == 0:
                try:
                    if final_json_output:
                        result = json.loads(final_json_output)
                        asyncio.run(event_queue.put({"type": "result", "data": result}))
                    else:
                        asyncio.run(event_queue.put({
                            "type": "error",
                            "message": "No output received from profile generation script"
                        }))
                except json.JSONDecodeError as e:
                    logger.error(f"Failed to parse profile generation output: {e}")
                    asyncio.run(event_queue.put({
                        "type": "error",
                        "message": f"Failed to parse profile generation output. Error: {str(e)}"
                    }))
            else:
                asyncio.run(event_queue.put({
                    "type": "error",
                    "message": f"Profile generation script failed with exit code {process.returncode}. Check server logs for details."
                }))

            # Signal completion
            asyncio.run(event_queue.put({"type": "done"}))

        except Exception as e:
            logger.error(f"Error in profile generation subprocess thread: {e}")
            logger.error(f"Exception type: {type(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            asyncio.run(event_queue.put({
                "type": "error",
                "message": f"Failed to start profile generation process: {str(e)}"
            }))
            asyncio.run(event_queue.put({"type": "done"}))

    # Start the subprocess in a separate thread
    subprocess_thread = threading.Thread(target=run_subprocess, daemon=True)
    subprocess_thread.start()


@router.get("/generation-status")
async def generation_status():
    """
    Server-Sent Events endpoint for real-time profile generation status.

    Streams progress events from the Python load profile generation script to the frontend.

    Returns:
        StreamingResponse: SSE stream with status events

    Event Types:
    - log: Progress log message
    - result: Final generation result
    - error: Error message
    - done: Process completed
    """
    global profile_event_queue

    async def event_generator():
        """Generate SSE events from the queue"""
        try:
            if profile_event_queue is None:
                raise HTTPException(status_code=500, detail="Event queue not initialized")

            while True:
                # Get event from queue
                event = await profile_event_queue.get()

                # Send event
                yield f"data: {json.dumps(event)}\n\n"

                # Check if this is the done event
                if event.get('type') == 'done':
                    break

        except Exception as e:
            logger.error(f"SSE error: {e}")
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )
