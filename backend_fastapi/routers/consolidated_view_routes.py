"""
Consolidated View Routes
========================

Consolidates electricity consumption data across multiple sectors and years.

Endpoints:
- POST /project/consolidated-electricity - Generate consolidated electricity view
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
import logging

logger = logging.getLogger(__name__)
router = APIRouter()

def find_cell_position(worksheet, marker: str) -> Optional[tuple]:
    """
    Find a marker cell (like '~Solar_share').

    Args:
        worksheet: openpyxl worksheet
        marker: Marker string to find

    Returns:
        Tuple of (row, col) if found, None otherwise
    """
    lower_marker = marker.lower()
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            if (isinstance(cell_value, str) and
                cell_value.strip().lower() == lower_marker):
                return (row_idx, col_idx)
    return None
class ConsolidatedElectricityRequest(BaseModel):
    """Request model for consolidated electricity view"""
    projectPath: str = Field(..., description="Project root path")
    sectorsOrder: Optional[List[str]] = Field(default=[], description="Ordered list of sectors")


@router.post("/consolidated-electricity")
async def consolidated_electricity(request: ConsolidatedElectricityRequest):
    """
    Consolidate electricity consumption by sectors and years.

    Reads all sheets from input_demand_file.xlsx that contain 'Year' and 'Electricity' columns,
    then consolidates the data into a single table.

    Args:
        request: Consolidation parameters

    Returns:
        dict: Consolidated data with years as rows and sectors as columns

    Raises:
        HTTPException: 400/404/500 on errors
    """
    project_path = request.projectPath
    sectors_order = request.sectorsOrder or []

    if not project_path:
        raise HTTPException(
            status_code=400,
            detail="Missing project path"
        )

    try:
        file_path = Path(project_path) / "inputs" / "input_demand_file.xlsx"

        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail="Excel file not found"
            )

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        year_wise = {}
        found_sectors = set()

        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Convert sheet to list of dicts
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                data.append(row_dict)

            # Check if sheet has Year and Electricity columns
            has_year = any('Year' in headers or 'year' in headers)
            has_electricity = any('Electricity' in headers or 'electricity' in headers)

            if not has_year or not has_electricity:
                continue

            sector = sheet_name.strip()
            found_sectors.add(sector)

            # Process each row
            for row in data:
                year = row.get('Year') or row.get('year')
                electricity = row.get('Electricity') or row.get('electricity')

                if not year or year == '':
                    continue

                try:
                    numeric_year = int(float(year))
                except (ValueError, TypeError):
                    continue

                if numeric_year not in year_wise:
                    year_wise[numeric_year] = {"Year": numeric_year}

                year_wise[numeric_year][sector] = electricity

        workbook.close()

        # Maintain only valid sectors in the order received
        ordered_sectors = [s for s in sectors_order if s in found_sectors]

        # Format the output
        formatted_array = []
        for year in sorted(year_wise.keys()):
            result_row = {"Year": year}
            for sector in ordered_sectors:
                result_row[sector] = year_wise[year].get(sector, '')
            formatted_array.append(result_row)

        return {"data": formatted_array}

    except HTTPException:
        raise
    except Exception as err:
        logger.error(f"❌ Error in consolidated-electricity API: {err}")
        raise HTTPException(
            status_code=500,
            detail="Internal Server Error"
        )
