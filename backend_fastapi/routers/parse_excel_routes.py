"""
Excel Parsing Routes
====================

Handles extraction of sector-specific data combined with economic indicators.

Endpoints:
- POST /project/extract-sector-data - Extract sector data with econometric parameters
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


class ExtractSectorDataRequest(BaseModel):
    """Request model for extracting sector data"""
    projectPath: str = Field(..., description="Project root path")
    sectorName: str = Field(..., description="Sector name to extract")


def find_sheet(workbook, sheet_name: str):
    """
    Find a sheet by case-insensitive name.

    Args:
        workbook: openpyxl workbook object
        sheet_name: Sheet name to find

    Returns:
        Worksheet if found, None otherwise
    """
    lower_case_name = sheet_name.lower()
    for name in workbook.sheetnames:
        if name.lower() == lower_case_name:
            return workbook[name]
    return None


def find_cell_position(worksheet, marker: str) -> Optional[tuple]:
    """
    Find a marker cell (like '~Consumption_Sectors' or '~Econometric_Parameters').

    Args:
        worksheet: openpyxl worksheet
        marker: Marker string to find

    Returns:
        Tuple of (row, col) if found, None otherwise
    """
    lower_marker = marker.lower()
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            if (isinstance(cell_value, str) and
                cell_value.strip().lower() == lower_marker):
                return (row_idx, col_idx)
    return None


@router.post("/extract-sector-data")
async def extract_sector_data(request: ExtractSectorDataRequest):
    """
    Extract sector-specific data merged with economic indicators.

    Process:
    1. Read econometric parameters for the sector from 'main' sheet
    2. Extract Year and Electricity data from sector-specific sheet
    3. Merge with economic indicator values from 'Economic_Indicators' sheet

    Args:
        request: Sector extraction parameters

    Returns:
        dict: Merged data with Year, Electricity, and economic indicators

    Raises:
        HTTPException: 400/404/500 on errors
    """
    project_path = request.projectPath
    sector_name = request.sectorName

    if not project_path or not sector_name:
        raise HTTPException(
            status_code=400,
            detail="Missing project path or sector name"
        )

    try:
        file_path = Path(project_path) / "inputs" / "input_demand_file.xlsx"

        if not file_path.exists():
            logger.error(f"[extract-sector-data] Excel file not found at path: {file_path}")
            raise HTTPException(
                status_code=404,
                detail="Excel file not found"
            )

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=True)

        # Find sheets (case-insensitive)
        main_sheet = find_sheet(workbook, 'main')
        econ_sheet = find_sheet(workbook, 'Economic_Indicators')

        if not main_sheet:
            logger.error("[extract-sector-data] Sheet 'main' not found in the workbook.")
            raise HTTPException(
                status_code=404,
                detail="Sheet 'main' not found. Please ensure it exists."
            )

        if not econ_sheet:
            logger.error("[extract-sector-data] Sheet 'Economic_Indicators' not found in the workbook.")
            raise HTTPException(
                status_code=404,
                detail="Sheet 'Economic_Indicators' not found. Please ensure it exists."
            )

        # 1. Get sector-specific economic parameters from Main sheet
        econ_param_marker = find_cell_position(main_sheet, '~Econometric_Parameters')
        if not econ_param_marker:
            logger.error("[extract-sector-data] Marker '~Econometric_Parameters' not found in 'main' sheet.")
            raise HTTPException(
                status_code=404,
                detail="Econometric marker not found"
            )

        marker_row, marker_col = econ_param_marker
        headers_row = marker_row + 1

        # Find the sector column
        sector_column = None
        max_col = main_sheet.max_column

        for col in range(marker_col, max_col + 1):
            cell_value = main_sheet.cell(row=headers_row, column=col).value
            if cell_value and str(cell_value).strip().lower() == sector_name.strip().lower():
                sector_column = col
                break

        if sector_column is None:
            logger.error(f"[extract-sector-data] Sector column '{sector_name}' not found under econometric parameters.")
            raise HTTPException(
                status_code=404,
                detail=f"Sector '{sector_name}' not found under econometric parameters"
            )

        # 2. Extract economic indicator names below this column
        indicators = []
        for row in range(headers_row + 1, main_sheet.max_row + 1):
            cell_value = main_sheet.cell(row=row, column=sector_column).value
            if not cell_value:
                break
            indicators.append(cell_value)

        # 3. Read sector sheet for Year & Electricity
        sector_sheet = find_sheet(workbook, sector_name)
        if not sector_sheet:
            logger.error(f"[extract-sector-data] Sector data sheet for '{sector_name}' not found.")
            raise HTTPException(
                status_code=404,
                detail=f"Sector sheet for '{sector_name}' not found"
            )

        # Convert sector sheet to list of dicts
        sector_data = []
        headers = [cell.value for cell in next(sector_sheet.iter_rows(min_row=1, max_row=1))]
        for row in sector_sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            sector_data.append(row_dict)

        # 4. Collect economic values from Economic_Indicators sheet
        econ_data = []
        econ_headers = [cell.value for cell in next(econ_sheet.iter_rows(min_row=1, max_row=1))]
        for row in econ_sheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(econ_headers, row))
            econ_data.append(row_dict)

        # 5. Merge data
        merged = []
        for sector_row in sector_data:
            year = sector_row.get('Year') or sector_row.get('year')
            electricity = sector_row.get('Electricity') or sector_row.get('electricity')

            # Find matching economic data for this year
            econ_row = next(
                (e for e in econ_data if (e.get('Year') or e.get('year')) == year),
                {}
            )

            obj = {"Year": year, "Electricity": electricity}
            for key in indicators:
                obj[key] = econ_row.get(key, None)

            merged.append(obj)

        workbook.close()

        return {"data": merged}

    except HTTPException:
        raise
    except Exception as err:
        logger.error(f"❌ Excel parse error in /extract-sector-data: {err}")
        raise HTTPException(
            status_code=500,
            detail="Internal Server Error during Excel parsing"
        )
