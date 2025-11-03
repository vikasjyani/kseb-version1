"""
Time Series Data Routes
=======================

Handles hourly load profile time series data extraction with filtering.

Endpoints:
- GET /project/full-load-profile - Get hourly load data filtered by fiscal year/month/season
"""

from fastapi import APIRouter, HTTPException, Query
from pathlib import Path
from typing import Optional
import openpyxl
from openpyxl.utils import datetime as excel_datetime
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/full-load-profile")
async def get_full_load_profile(
    projectPath: str = Query(..., description="Project root path"),
    profileName: str = Query(..., description="Profile name"),
    fiscalYear: str = Query(..., description="Fiscal year (e.g., FY2025)"),
    month: Optional[int] = Query(None, description="Month number (1-12)"),
    season: Optional[str] = Query(None, description="Season name")
):
    """
    Get full hourly load profile data with optional filtering by month or season.

    Args:
        projectPath: Project root directory
        profileName: Name of the profile (without .xlsx)
        fiscalYear: Fiscal year string (e.g., 'FY2025')
        month: Optional month filter (1-12)
        season: Optional season filter (Monsoon, Post-monsoon, Winter, Summer)

    Returns:
        dict: Filtered hourly load profile data
    """
    if not projectPath or not profileName or not fiscalYear:
        raise HTTPException(
            status_code=400,
            detail="Project path, profile name, and fiscal year are required."
        )

    # Parse fiscal year
    try:
        year_to_filter = int(fiscalYear.replace('FY', ''))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid fiscal year format.")

    # Define season-to-month mapping
    season_months = {
        'Monsoon': [7, 8, 9],
        'Post-monsoon': [10, 11],
        'Winter': [12, 1, 2],
        'Summer': [3, 4, 5, 6]
    }

    # Validate and prepare month filter
    months_to_filter = []
    if month:
        if not (1 <= month <= 12):
            raise HTTPException(
                status_code=400,
                detail="Invalid month. Must be a number between 1 and 12."
            )
        months_to_filter = [month]
    elif season:
        if season not in season_months:
            raise HTTPException(
                status_code=400,
                detail="Invalid season. Must be one of: Monsoon, Post-monsoon, Winter, Summer."
            )
        months_to_filter = season_months[season]

    file_path = Path(projectPath) / "results" / "load_profiles" / f"{profileName}.xlsx"

    try:
        if not file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Profile file not found: {profileName}.xlsx"
            )

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = 'Load_Profile'

        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found.")

        worksheet = workbook[sheet_name]

        # Read all data
        headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        all_data = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            all_data.append(row_dict)

        workbook.close()

        # Filter data by fiscal year
        filtered_data = [
            row for row in all_data
            if row.get('Fiscal_Year') == year_to_filter
        ]

        # Filter by month(s) if specified
        if months_to_filter:
            filtered_data = [
                row for row in filtered_data
                if row.get('Month') in months_to_filter
            ]

        # Format DateTime field
        for row in filtered_data:
            excel_date_serial = row.get('DateTime')
            if excel_date_serial and isinstance(excel_date_serial, (int, float)):
                # Convert Excel serial number to datetime
                try:
                    from datetime import datetime, timedelta
                    # Excel epoch starts at 1899-12-30
                    excel_epoch = datetime(1899, 12, 30)
                    python_date = excel_epoch + timedelta(days=excel_date_serial)
                    row['DateTime'] = python_date.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    # If conversion fails, keep original value
                    pass

        return {"success": True, "data": filtered_data}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"❌ Error reading full load profile for '{profileName}': {error}")
        raise HTTPException(
            status_code=500,
            detail="An error occurred while reading the profile file."
        )
