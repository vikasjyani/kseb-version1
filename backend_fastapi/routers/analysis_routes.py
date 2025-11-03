"""
Analysis Routes
===============

Handles load profile analysis data extraction.

Endpoints:
- GET /project/analysis-data - Get monthly/seasonal analysis from Excel
- GET /project/profile-years - List fiscal years in profile
"""

from fastapi import APIRouter, HTTPException, Query
from pathlib import Path
from typing import List
import openpyxl
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/analysis-data")
async def get_analysis_data(
    projectPath: str = Query(..., description="Project root path"),
    profileName: str = Query(..., description="Profile name"),
    sheetName: str = Query(..., description="Sheet name to read")
):
    """
    Get analysis data from a specific sheet in the load profile.

    Reads monthly or seasonal analysis data and structures it by parameter.

    Args:
        projectPath: Project root directory
        profileName: Name of the profile (without .xlsx)
        sheetName: Name of the sheet to read

    Returns:
        dict: Structured data grouped by parameter, with sorted columns
    """
    if not projectPath or not profileName or not sheetName:
        raise HTTPException(status_code=400, detail="Missing required parameters.")

    file_path = Path(projectPath) / "results" / "load_profiles" / f"{profileName}.xlsx"

    try:
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Profile file not found.")

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheetName not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(status_code=404, detail=f"Sheet '{sheetName}' not found.")

        worksheet = workbook[sheetName]

        # Read headers
        headers = [str(cell.value) for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

        if len(headers) == 0:
            workbook.close()
            return {"success": True, "data": {}, "columns": []}

        # Read data rows
        data_rows = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            data_rows.append(row_dict)

        workbook.close()

        # Filter columns to sort (exclude 'Parameters' and 'Fiscal_Year')
        columns_to_sort = [h for h in headers if h not in ['Parameters', 'Fiscal_Year']]

        # Define sort order for months and seasons
        sort_order_map = {
            # Months (number and name)
            '4': 1, 'April': 1,
            '5': 2, 'May': 2,
            '6': 3, 'June': 3,
            '7': 4, 'July': 4,
            '8': 5, 'August': 5,
            '9': 6, 'September': 6,
            '10': 7, 'October': 7,
            '11': 8, 'November': 8,
            '12': 9, 'December': 9,
            '1': 10, 'January': 10,
            '2': 11, 'February': 11,
            '3': 12, 'March': 12,
            # Seasons
            'Summer': 21,
            'Monsoon': 22,
            'Post-monsoon': 23,
            'Winter': 24,
        }

        # Sort columns
        columns_to_sort.sort(key=lambda x: (sort_order_map.get(x, 99), str(x)))

        # Structure data by parameter
        structured_data = {}
        for row in data_rows:
            parameter = row.get('Parameters')
            if parameter:
                if parameter not in structured_data:
                    structured_data[parameter] = []

                row_data = dict(row)
                del row_data['Parameters']
                structured_data[parameter].append(row_data)

        return {"success": True, "data": structured_data, "columns": columns_to_sort}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"❌ Error in /analysis-data for '{profileName}': {error}")
        raise HTTPException(status_code=500, detail="An internal server error occurred.")


@router.get("/profile-years")
async def get_profile_years(
    projectPath: str = Query(..., description="Project root path"),
    profileName: str = Query(..., description="Profile name")
):
    """
    Extract unique fiscal years from a load profile.

    Args:
        projectPath: Project root directory
        profileName: Name of the profile (without .xlsx)

    Returns:
        dict: List of fiscal years (e.g., ['FY2025', 'FY2026'])
    """
    if not projectPath or not profileName:
        raise HTTPException(status_code=400, detail="Project path and profile name are required.")

    file_path = Path(projectPath) / "results" / "load_profiles" / f"{profileName}.xlsx"

    try:
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Profile file not found.")

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = 'Load_Profile'

        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(status_code=404, detail="Sheet 'Load_Profile' not found.")

        # Read data
        headers = [cell.value for cell in next(workbook[sheet_name].iter_rows(min_row=1, max_row=1))]
        json_data = []
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            json_data.append(row_dict)

        workbook.close()

        # Extract unique years
        unique_years = sorted(set(
            row['Fiscal_Year'] for row in json_data
            if row.get('Fiscal_Year') is not None
        ))

        # Format as FY{year}
        formatted_years = [f"FY{int(year)}" for year in unique_years]

        return {"success": True, "years": formatted_years}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"❌ Error in /profile-years for '{profileName}': {error}")
        raise HTTPException(status_code=500, detail="An error occurred.")
