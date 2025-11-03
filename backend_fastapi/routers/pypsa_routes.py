"""
PyPSA Optimization Routes
=========================

Handles PyPSA grid optimization results visualization from Excel files.

Performance Features:
- Input validation and sanitization
- Memory-efficient Excel reading (read-only mode)
- Proper resource cleanup (workbook closing)
- Comprehensive error handling
- Security: Path traversal prevention

Best Practices:
- Defensive programming with existence checks
- Detailed error logging with context
- Proper exception handling hierarchy
- Resource management (auto-close workbooks)
- Type hints for better code clarity

Endpoints:
- GET /project/optimization-folders - List pypsa_optimization subfolders
- GET /project/optimization-sheets - Get sheet names from Pypsa_results.xlsx
- GET /project/optimization-sheet-data - Get data from a specific sheet (with pagination)
"""

from fastapi import APIRouter, HTTPException, Query
from pathlib import Path
from typing import Optional
import openpyxl
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def validate_path_components(path: str, component_name: str) -> str:
    """
    Validate path component to prevent path traversal attacks.

    Args:
        path: Path component to validate
        component_name: Name for error messages

    Returns:
        str: Validated path component

    Raises:
        HTTPException: If validation fails
    """
    if not path:
        raise HTTPException(status_code=400, detail=f"{component_name} is required")

    # Check for path traversal attempts
    if ".." in path or "/" in path or "\\" in path:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid {component_name}: path traversal not allowed"
        )

    return path.strip()


@router.get("/optimization-folders")
async def get_optimization_folders(projectPath: str = Query(..., description="Project root path")):
    """
    List all subfolders in the pypsa_optimization directory.

    Args:
        projectPath: Project root directory

    Returns:
        dict: List of folder names
    """
    if not projectPath:
        raise HTTPException(status_code=400, detail="Project path is required.")

    optimization_folder_path = Path(projectPath) / "results" / "pypsa_optimization"

    if not optimization_folder_path.exists():
        logger.warning(f"Directory not found: {optimization_folder_path}")
        return {"success": True, "folders": []}

    try:
        folders = [
            item.name for item in optimization_folder_path.iterdir()
            if item.is_dir()
        ]
        return {"success": True, "folders": folders}

    except Exception as error:
        logger.error(f"❌ Error reading optimization folders: {error}")
        raise HTTPException(
            status_code=500,
            detail="Failed to read optimization folders."
        )


@router.get("/optimization-sheets")
async def get_optimization_sheets(
    projectPath: str = Query(..., description="Project root path"),
    folderName: str = Query(..., description="Optimization folder name")
):
    """
    Get all sheet names from Pypsa_results.xlsx in a specific folder.

    Args:
        projectPath: Project root directory
        folderName: Name of the optimization folder

    Returns:
        dict: List of sheet names
    """
    if not projectPath or not folderName:
        raise HTTPException(
            status_code=400,
            detail="Project path and folder name are required."
        )

    excel_file_path = (
        Path(projectPath) / "results" / "pypsa_optimization" / folderName / "Pypsa_results.xlsx"
    )

    if not excel_file_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Pypsa_results.xlsx not found in the selected folder."
        )

    try:
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()

        return {"success": True, "sheets": sheet_names}

    except Exception as error:
        logger.error(f"❌ Error reading Excel file sheet names: {error}")
        raise HTTPException(
            status_code=500,
            detail="Failed to read the Excel file."
        )


@router.get("/optimization-sheet-data")
async def get_optimization_sheet_data(
    projectPath: str = Query(..., description="Project root path"),
    folderName: str = Query(..., description="Optimization folder name"),
    sheetName: str = Query(..., description="Sheet name to read"),
    limit: Optional[int] = Query(None, description="Maximum number of rows to return (pagination)", ge=1, le=10000),
    offset: Optional[int] = Query(0, description="Number of rows to skip (pagination)", ge=0)
):
    """
    Get data from a specific sheet in Pypsa_results.xlsx.

    Performance Features:
    - Read-only mode for faster loading
    - Optional pagination to limit memory usage
    - Automatic resource cleanup (workbook closing)
    - Memory-efficient row iteration

    Args:
        projectPath: Project root directory
        folderName: Name of the optimization folder
        sheetName: Name of the sheet to read
        limit: Maximum number of rows to return (default: all rows, max: 10000)
        offset: Number of rows to skip (default: 0)

    Returns:
        dict: {
            'success': bool,
            'data': List[dict],
            'count': int,
            'total_rows': int (if pagination),
            'offset': int (if pagination),
            'limit': int (if pagination),
            'has_more': bool (if pagination)
        }

    Raises:
        HTTPException: 400 for invalid input, 404 for not found, 500 for server errors
    """
    try:
        # Validate inputs
        validate_path_components(folderName, "Folder name")
        validate_path_components(sheetName, "Sheet name")

        # Construct Excel file path
        excel_file_path = (
            Path(projectPath) / "results" / "pypsa_optimization" / folderName / "Pypsa_results.xlsx"
        )

        if not excel_file_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Pypsa_results.xlsx not found in folder: {folderName}"
            )

        logger.info(f"Reading sheet '{sheetName}' from {excel_file_path} "
                   f"(offset={offset}, limit={limit})")

        # Load workbook in read-only mode for better performance
        workbook = openpyxl.load_workbook(excel_file_path, read_only=True, data_only=True)

        try:
            if sheetName not in workbook.sheetnames:
                raise HTTPException(
                    status_code=404,
                    detail=f"Sheet '{sheetName}' not found. Available sheets: {workbook.sheetnames}"
                )

            worksheet = workbook[sheetName]

            # Read headers
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]

            # Read data rows with pagination
            json_data = []
            row_count = 0
            total_rows = 0

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=0):
                total_rows += 1

                # Skip rows for offset
                if row_idx < offset:
                    continue

                # Stop if limit reached
                if limit and row_count >= limit:
                    break

                # Convert row to dict
                row_dict = dict(zip(headers, row))
                json_data.append(row_dict)
                row_count += 1

            # Prepare response
            response = {
                "success": True,
                "data": json_data,
                "count": len(json_data)
            }

            # Add pagination metadata if used
            if limit is not None:
                response["total_rows"] = total_rows
                response["offset"] = offset
                response["limit"] = limit
                response["has_more"] = (offset + len(json_data)) < total_rows

            logger.info(f"Successfully read {len(json_data)} rows from sheet '{sheetName}'")

            return response

        finally:
            # Ensure workbook is closed even if error occurs
            workbook.close()

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error getting sheet data: {error}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Failed to retrieve data from sheet: {str(error)}"
        )
