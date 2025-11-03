"""
Scenario Management Routes
==========================

Handles demand forecast scenario operations including T&D losses and consolidated results.

Endpoints:
- GET /project/scenarios - List all scenario folders
- GET /project/scenarios/{scenarioName}/meta - Get scenario metadata
- GET /project/scenarios/{scenarioName}/sectors - List sectors in scenario
- GET /project/scenarios/{scenarioName}/models - Get available models by sector
- GET /project/scenarios/{scenarioName}/sectors/{sectorName} - Get sector forecast data
- GET /project/scenarios/{scenarioName}/td-losses - Get T&D loss configuration
- POST /project/scenarios/{scenarioName}/td-losses - Save T&D loss configuration
- GET /project/scenarios/{scenarioName}/consolidated/exists - Check if consolidated file exists
- POST /project/scenarios/{scenarioName}/consolidated - Generate consolidated results
- POST /project/save-consolidated - Save consolidated results to Excel
"""

from fastapi import APIRouter, HTTPException, Query, Path as PathParam
from pydantic import BaseModel, Field
from pathlib import Path
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.utils import get_column_letter
import json
import logging

logger = logging.getLogger(__name__)
router = APIRouter()


class TDLossPoint(BaseModel):
    """T&D loss data point"""
    year: int
    loss: float


class TDLossSaveRequest(BaseModel):
    """Request model for saving T&D losses"""
    projectPath: str
    lossPoints: List[TDLossPoint]


class ConsolidatedRequest(BaseModel):
    """Request model for generating consolidated results"""
    projectPath: str
    startYear: int
    endYear: int
    selections: Optional[Dict[str, str]] = {}
    demandType: Optional[str] = "gross"  # Options: "gross", "net", "onGrid"


class SaveConsolidatedRequest(BaseModel):
    """Request model for saving consolidated results"""
    projectPath: str
    scenarioName: str
    data: List[Dict[str, Any]]


def read_solar_share_data(project_path: str) -> Dict[str, float]:
    """
    Read solar share percentages for each sector from input_demand_file.xlsx.

    Looks for the ~Solar_share marker in the 'main' sheet and reads
    Sector and Percentage_share columns below it.

    Args:
        project_path: Project root directory path

    Returns:
        Dictionary mapping sector name to percentage share (e.g., {"Agriculture": 5.5})
    """
    try:
        file_path = Path(project_path) / "inputs" / "input_demand_file.xlsx"

        if not file_path.exists():
            logger.warning(f"[read_solar_share_data] Excel file not found at: {file_path}")
            return {}

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Find 'main' sheet (case-insensitive)
        main_sheet = None
        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == 'main':
                main_sheet = workbook[sheet_name]
                break

        if not main_sheet:
            logger.warning("[read_solar_share_data] Sheet 'main' not found")
            workbook.close()
            return {}

        # Find ~Solar_share marker
        marker_row = None
        marker_col = None
        for row_idx, row in enumerate(main_sheet.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if (isinstance(cell_value, str) and
                    cell_value.strip().lower() == '~solar_share'):
                    marker_row = row_idx
                    marker_col = col_idx
                    break
            if marker_row:
                break

        if not marker_row:
            logger.info("[read_solar_share_data] Marker '~Solar_share' not found, returning empty dict")
            workbook.close()
            return {}

        # Read header row (should be 1 or 2 rows below marker)
        # Try row immediately after marker
        headers_row = marker_row + 1
        headers = [main_sheet.cell(row=headers_row, column=col).value for col in range(1, main_sheet.max_column + 1)]

        # If first row after marker is blank, try next row
        if not any(headers):
            headers_row = marker_row + 2
            headers = [main_sheet.cell(row=headers_row, column=col).value for col in range(1, main_sheet.max_column + 1)]

        # Find Sector and Percentage_share columns
        sector_col = None
        percentage_col = None

        for col_idx, header in enumerate(headers, start=1):
            if header:
                header_lower = str(header).strip().lower()
                if header_lower == 'sector':
                    sector_col = col_idx
                elif 'percentage' in header_lower and 'share' in header_lower:
                    percentage_col = col_idx

        if not sector_col or not percentage_col:
            logger.warning(f"[read_solar_share_data] Required columns not found. Sector col: {sector_col}, Percentage col: {percentage_col}")
            workbook.close()
            return {}

        # Read data rows
        solar_shares = {}
        for row_idx in range(headers_row + 1, main_sheet.max_row + 1):
            sector_name = main_sheet.cell(row=row_idx, column=sector_col).value
            percentage_value = main_sheet.cell(row=row_idx, column=percentage_col).value

            if not sector_name or sector_name == '':
                break  # Stop at first empty sector

            try:
                percentage_float = float(percentage_value) if percentage_value else 0.0
                solar_shares[str(sector_name).strip()] = percentage_float
            except (ValueError, TypeError):
                logger.warning(f"[read_solar_share_data] Invalid percentage value for sector {sector_name}: {percentage_value}")
                solar_shares[str(sector_name).strip()] = 0.0

        workbook.close()
        logger.info(f"[read_solar_share_data] Successfully loaded solar shares for {len(solar_shares)} sectors")
        return solar_shares

    except Exception as e:
        logger.error(f"[read_solar_share_data] Error reading solar share data: {e}")
        return {}


def is_solar_sector(sector_name: str) -> bool:
    """
    Check if a sector name contains 'solar' or 'Solar'.

    Args:
        sector_name: Name of the sector

    Returns:
        True if sector is a solar generation sector
    """
    return 'solar' in sector_name.lower()


def calculate_td_loss_percentage(target_year: int, loss_points: List[TDLossPoint]) -> float:
    """
    Calculate T&D loss percentage for a given year using linear interpolation.

    Args:
        target_year: Year to calculate loss for
        loss_points: List of T&D loss data points

    Returns:
        Loss percentage as decimal (e.g., 0.10 for 10%)
    """
    DEFAULT_LOSS = 0.10  # 10%

    if not loss_points or len(loss_points) == 0:
        return DEFAULT_LOSS

    # Filter and sort points
    sorted_points = sorted(
        [p for p in loss_points if isinstance(p.year, (int, float)) and isinstance(p.loss, (int, float))],
        key=lambda p: p.year
    )

    if len(sorted_points) == 0:
        return DEFAULT_LOSS
    if len(sorted_points) == 1:
        return sorted_points[0].loss / 100

    first_point = sorted_points[0]
    last_point = sorted_points[-1]

    # Extrapolation
    if target_year <= first_point.year:
        return first_point.loss / 100
    if target_year >= last_point.year:
        return last_point.loss / 100

    # Interpolation
    for i in range(len(sorted_points) - 1):
        p1 = sorted_points[i]
        p2 = sorted_points[i + 1]

        if target_year >= p1.year and target_year <= p2.year:
            if p2.year - p1.year == 0:
                return p1.loss / 100

            # Linear interpolation
            interpolated_loss = p1.loss + (target_year - p1.year) * (p2.loss - p1.loss) / (p2.year - p1.year)
            return interpolated_loss / 100

    return DEFAULT_LOSS


@router.get("/scenarios")
async def list_scenarios(projectPath: str = Query(..., description="Project root path")):
    """
    List all demand forecast scenario folders.

    Args:
        projectPath: Project root directory

    Returns:
        dict: List of scenario names
    """
    if not projectPath:
        raise HTTPException(status_code=400, detail="Missing projectPath")

    scenario_base_path = Path(projectPath) / "results" / "demand_forecasts"

    if not scenario_base_path.exists():
        return {"scenarios": []}

    try:
        scenarios = [
            item.name for item in scenario_base_path.iterdir()
            if item.is_dir()
        ]
        return {"scenarios": scenarios}
    except Exception as e:
        logger.error(f"Error reading scenario folders: {e}")
        raise HTTPException(status_code=500, detail="Failed to read scenario folders")


@router.get("/scenarios/{scenarioName}/meta")
async def get_scenario_meta(
    scenarioName: str = PathParam(..., description="Scenario name"),
    projectPath: str = Query(..., description="Project root path")
):
    """
    Get scenario metadata from scenario_meta.json file.

    Args:
        scenarioName: Name of the scenario
        projectPath: Project root directory

    Returns:
        dict: Scenario metadata
    """
    if not projectPath or not scenarioName:
        raise HTTPException(status_code=400, detail="Project path and scenario name are required.")

    try:
        meta_file_path = (
            Path(projectPath) / "results" / "demand_forecasts" / scenarioName / "scenario_meta.json"
        )

        if meta_file_path.exists():
            with open(meta_file_path, 'r') as f:
                meta_data = json.load(f)
            return {"success": True, "meta": meta_data}
        else:
            raise HTTPException(status_code=404, detail="Scenario metadata not found.")

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error reading metadata for {scenarioName}: {error}")
        raise HTTPException(status_code=500, detail="Failed to read scenario metadata.")


@router.get("/scenarios/{scenarioName}/sectors")
async def get_scenario_sectors(
    scenarioName: str = PathParam(..., description="Scenario name"),
    projectPath: str = Query(..., description="Project root path")
):
    """
    List all sector Excel files in a scenario.

    Args:
        scenarioName: Name of the scenario
        projectPath: Project root directory

    Returns:
        dict: List of sector names (Excel files without extension)
    """
    if not projectPath or not scenarioName:
        raise HTTPException(status_code=400, detail="Project path and scenario name are required.")

    try:
        scenario_path = Path(projectPath) / "results" / "demand_forecasts" / scenarioName

        if not scenario_path.exists():
            raise HTTPException(status_code=404, detail="Scenario folder not found.")

        files = [f.stem for f in scenario_path.glob("*.xlsx")]
        return {"success": True, "sectors": files}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error fetching sectors: {error}")
        raise HTTPException(status_code=500, detail="Failed to fetch sectors.")


@router.get("/scenarios/{scenarioName}/models")
async def get_scenario_models(
    scenarioName: str = PathParam(..., description="Scenario name"),
    projectPath: str = Query(..., description="Project root path")
):
    """
    Get available forecasting models for each sector.

    Reads the 'Results' sheet headers from each sector Excel file.

    Args:
        scenarioName: Name of the scenario
        projectPath: Project root directory

    Returns:
        dict: Dictionary mapping sector names to list of model names
    """
    if not projectPath or not scenarioName:
        raise HTTPException(status_code=400, detail="Project path and scenario name are required.")

    try:
        scenario_path = Path(projectPath) / "results" / "demand_forecasts" / scenarioName

        if not scenario_path.exists():
            raise HTTPException(status_code=404, detail="Scenario folder not found.")

        models_by_sector = {}

        for file in scenario_path.glob("*.xlsx"):
            sector_name = file.stem
            if sector_name == 'Consolidated_Results':
                continue

            try:
                workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
                if 'Results' in workbook.sheetnames:
                    sheet = workbook['Results']
                    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                    # Exclude 'Year' and 'Time Series' columns
                    models = [
                        h for h in headers
                        if h and h.lower() not in ['year', 'time series']
                    ]
                    models_by_sector[sector_name] = models
                else:
                    models_by_sector[sector_name] = []
                workbook.close()
            except Exception as e:
                logger.error(f"Error reading {file}: {e}")
                models_by_sector[sector_name] = []

        return {"success": True, "models": models_by_sector}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error fetching models: {error}")
        raise HTTPException(status_code=500, detail="Failed to fetch models for the scenario.")

from pathlib import Path
import openpyxl
from fastapi import HTTPException
# PathParam and Query typed-deps assumed imported as in your original file
# logger assumed to be available in module scope

@router.get("/scenarios/{scenarioName}/sectors/{sectorName}")
async def get_sector_data(
    scenarioName: str = PathParam(..., description="Scenario name"),
    sectorName: str = PathParam(..., description="Sector name"),
    projectPath: str = Query(..., description="Project root path"),
    startYear: int = Query(..., description="Start year filter"),
    endYear: int = Query(..., description="End year filter")
    ):
    """
    Get forecast data for a specific sector within a year range.

    Returns:
        dict: Filtered forecast data with forecastYearstart (int or None)
    """
    # basic validation
    if not all([projectPath, scenarioName, sectorName]) or startYear is None or endYear is None:
        raise HTTPException(status_code=400, detail="Missing required parameters.")
    if startYear > endYear:
        raise HTTPException(status_code=400, detail="startYear must be <= endYear.")

    try:
        file_path = Path(projectPath) / "results" / "demand_forecasts" / scenarioName / f"{sectorName}.xlsx"

        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Sector result file not found.")

        # choose sheet name for result rows
        sheet_name = "Consolidated Data" if sectorName == "Consolidated_Results" else "Results"

        # Open workbook once (read_only for memory efficiency)
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise HTTPException(status_code=404, detail=f"The '{sheet_name}' sheet was not found in the Excel file.")
        result_sheet = workbook[sheet_name]

        # Compute forecast_year_start from Inputs sheet (only for non-consolidated sectors)
        forecast_year_start = None
        if sectorName != "Consolidated_Results":
            try:
                if "Inputs" in workbook.sheetnames:
                    inputs_sheet = workbook["Inputs"]

                    # read header row for Inputs
                    inputs_iter = inputs_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
                    try:
                        inputs_headers = next(inputs_iter)
                    except StopIteration:
                        inputs_headers = []

                    if inputs_headers:
                        # Normalize headers to strings
                        headers = [h if h is not None else "" for h in inputs_headers]
                        # find column indices for 'Year' and 'Electricity' (case-sensitive match as before)
                        try:
                            year_idx = headers.index("Year")
                            elec_idx = headers.index("Electricity")
                        except ValueError:
                            year_idx = elec_idx = None

                        max_year = None
                        if year_idx is not None and elec_idx is not None:
                            # iterate remaining rows
                            for row in inputs_sheet.iter_rows(min_row=2, values_only=True):
                                # guard for short rows
                                if row is None:
                                    continue
                                # get electricity value for row
                                elec_val = row[elec_idx] if elec_idx < len(row) else None
                                if elec_val is None:
                                    continue  # skip rows with no Electricity
                                # get year value
                                year_val = row[year_idx] if year_idx < len(row) else None
                                if year_val is None:
                                    continue
                                # coerce to int if possible
                                try:
                                    # allow float-like years (e.g., 2020.0)
                                    year_int = int(year_val)
                                except (ValueError, TypeError):
                                    # if it's a string with whitespace, try stripping then int
                                    try:
                                        year_int = int(str(year_val).strip())
                                    except Exception:
                                        continue
                                if max_year is None or year_int > max_year:
                                    max_year = year_int
                            if max_year is not None:
                                forecast_year_start = int(max_year)
                            else:
                                # no valid rows found
                                forecast_year_start = None
                        else:
                            # missing columns
                            logger.warning(f"[get_sector_data] Missing 'Year' or 'Electricity' columns in Inputs sheet for {sectorName}")
                            forecast_year_start = None
                    else:
                        logger.warning(f"[get_sector_data] Inputs sheet header missing or empty for {sectorName}")
                        forecast_year_start = None
                else:
                    logger.warning(f"[get_sector_data] Inputs sheet not present for {sectorName}")
                    forecast_year_start = None
            except Exception as e:
                logger.warning(f"[get_sector_data] Could not determine forecast_year_start for {sectorName}: {e}")
                forecast_year_start = None

        # Read header row from result sheet
        result_headers_iter = result_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            result_headers = next(result_headers_iter)
            if result_headers is None:
                result_headers = []
        except StopIteration:
            result_headers = []

        headers = [h if h is not None else "" for h in result_headers]

        # Collect data rows
        data = []
        for row in result_sheet.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            # build dict safely even if row is shorter/longer than headers
            row_dict = {}
            for idx, header in enumerate(headers):
                # skip empty header names
                if header == "":
                    continue
                value = row[idx] if idx < len(row) else None
                row_dict[header] = value
            data.append(row_dict)

        # close workbook now that we're done reading
        workbook.close()

        # Filter by year range (coerce Year to int safely)
        filtered_data = []
        for row in data:
            year_val = row.get("Year")
            if year_val is None:
                continue
            try:
                year_int = int(year_val)
            except (ValueError, TypeError):
                # try stripping string and converting
                try:
                    year_int = int(str(year_val).strip())
                except Exception:
                    continue
            if startYear <= year_int <= endYear:
                filtered_data.append(row)

        return {
            "success": True,
            "data": filtered_data,
            "forecastYearstart": forecast_year_start
        }

    except HTTPException:
        # re-raise known fastapi exceptions
        raise
    except Exception as error:
        logger.error(f"Error processing sector data for {scenarioName}/{sectorName}: {error}", exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to process sector data.")

@router.get("/scenarios/{scenarioName}/td-losses")
async def get_td_losses(
    scenarioName: str = PathParam(..., description="Scenario name"),
    projectPath: str = Query(..., description="Project root path")
):
    """
    Get T&D loss configuration for a scenario.

    Args:
        scenarioName: Name of the scenario
        projectPath: Project root directory

    Returns:
        dict: T&D loss data points
    """
    if not projectPath or not scenarioName:
        raise HTTPException(status_code=400, detail="Missing required parameters.")

    try:
        file_path = (
            Path(projectPath) / "results" / "demand_forecasts" / scenarioName / "td_losses.json"
        )

        if file_path.exists():
            with open(file_path, 'r') as f:
                data = json.load(f)
            return {"success": True, "data": data}

        return {"success": True, "data": []}

    except Exception as error:
        logger.error(f"Error loading T&D loss data: {error}")
        raise HTTPException(status_code=500, detail="Failed to load T&D loss data.")


@router.post("/scenarios/{scenarioName}/td-losses")
async def save_td_losses(
    scenarioName: str = PathParam(..., description="Scenario name"),
    request: TDLossSaveRequest = None
):
    """
    Save T&D loss configuration for a scenario.

    Args:
        scenarioName: Name of the scenario
        request: T&D loss data points

    Returns:
        dict: Success message
    """
    if not request.projectPath or not scenarioName or request.lossPoints is None:
        raise HTTPException(status_code=400, detail="Missing required parameters.")

    try:
        scenario_path = Path(request.projectPath) / "results" / "demand_forecasts" / scenarioName
        scenario_path.mkdir(parents=True, exist_ok=True)

        file_path = scenario_path / "td_losses.json"

        # Convert Pydantic models to dicts
        loss_points_data = [point.dict() for point in request.lossPoints]

        with open(file_path, 'w') as f:
            json.dump(loss_points_data, f, indent=2)

        return {"success": True, "message": "T&D losses saved successfully."}

    except Exception as error:
        logger.error(f"Error saving T&D loss data: {error}")
        raise HTTPException(status_code=500, detail="Failed to save T&D loss data.")


@router.get("/scenarios/{scenarioName}/consolidated/exists")
async def check_consolidated_exists(
    scenarioName: str = PathParam(..., description="Scenario name"),
    projectPath: str = Query(..., description="Project root path")
):
    """
    Check if consolidated results file exists.

    Args:
        scenarioName: Name of the scenario
        projectPath: Project root directory

    Returns:
        dict: Boolean indicating file existence
    """
    if not projectPath or not scenarioName:
        raise HTTPException(status_code=400, detail="Project path and scenario name are required.")

    try:
        file_path = (
            Path(projectPath) / "results" / "demand_forecasts" / scenarioName / "Consolidated_Results.xlsx"
        )
        exists = file_path.exists()
        return {"success": True, "exists": exists}

    except Exception as error:
        logger.error(f"Error checking file existence: {error}")
        raise HTTPException(status_code=500, detail="Error checking file existence.")


@router.post("/scenarios/{scenarioName}/consolidated")
async def generate_consolidated(
    scenarioName: str = PathParam(..., description="Scenario name"),
    request: ConsolidatedRequest = None
):
    """
    Generate consolidated forecast results from individual sector files.

    Combines selected models from each sector and applies T&D losses.

    Args:
        scenarioName: Name of the scenario
        request: Consolidation parameters

    Returns:
        dict: Consolidated data with T&D losses applied
    """
    if not request.projectPath or not scenarioName or not request.startYear or not request.endYear:
        raise HTTPException(status_code=400, detail="Missing required parameters.")

    try:
        scenario_path = Path(request.projectPath) / "results" / "demand_forecasts" / scenarioName

        if not scenario_path.exists():
            raise HTTPException(status_code=404, detail="Scenario folder not found.")

        model_selections = request.selections or {}

        # Get all Excel files except Consolidated_Results
        excel_files = [
            f for f in scenario_path.glob("*.xlsx")
            if f.stem != 'Consolidated_Results'
        ]
        sectors = [f.stem for f in excel_files]

        # Create year range
        years = list(range(request.startYear, request.endYear + 1))

        # Initialize consolidated data
        consolidated_data = []
        data_map = {}
        for year in years:
            row_data = {"Year": year}
            for sector in sectors:
                row_data[sector] = None
            consolidated_data.append(row_data)
            data_map[year] = row_data

        # Read data from each sector
        for sector in sectors:
            selected_model = model_selections.get(sector)
            if not selected_model:
                continue

            file_path = scenario_path / f"{sector}.xlsx"
            if not file_path.exists():
                continue

            try:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if 'Results' not in workbook.sheetnames:
                    workbook.close()
                    continue

                sheet = workbook['Results']
                headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    year = row_dict.get('Year')
                    if year and int(year) in data_map:
                        value = row_dict.get(selected_model)
                        if value is not None:
                            data_map[int(year)][sector] = value

                workbook.close()

            except Exception as e:
                logger.error(f"Error reading sector {sector}: {e}")
                continue

        # Load solar share data for net demand calculation
        solar_shares = {}
        demand_type = request.demandType or "gross"

        if demand_type in ["net", "onGrid"]:
            solar_shares = read_solar_share_data(request.projectPath)
            logger.info(f"[generate_consolidated] Loaded solar shares: {solar_shares}")

        # Load T&D losses
        td_loss_points = []
        losses_file_path = scenario_path / "td_losses.json"
        if losses_file_path.exists():
            try:
                with open(losses_file_path, 'r') as f:
                    loss_data = json.load(f)
                    td_loss_points = [TDLossPoint(**point) for point in loss_data]
            except Exception as e:
                logger.error(f"Could not parse td_losses.json: {e}")

        # Calculate totals based on demand type
        for row in consolidated_data:
            year = row['Year']

            # Apply solar sector abs() transformation
            for sector in sectors:
                if row[sector] is not None and is_solar_sector(sector):
                    row[sector] = abs(float(row[sector]))

            if demand_type == "gross":
                # GROSS DEMAND: Original behavior
                gross_total = sum(
                    float(row[sector]) if row[sector] is not None else 0
                    for sector in sectors
                )

                td_percentage = calculate_td_loss_percentage(year, td_loss_points)
                td_losses = gross_total * (td_percentage / (1 - td_percentage))
                final_total = gross_total + td_losses

                row['Gross Total'] = gross_total
                row['T&D Loss (%)'] = td_percentage
                row['T&D Losses'] = td_losses
                row['Total'] = final_total

            elif demand_type == "net":
                # NET DEMAND: Subtract solar generation from each sector
                net_total = 0

                for sector in sectors:
                    if row[sector] is not None:
                        sector_value = float(row[sector])

                        # For non-solar sectors, subtract their solar share
                        if not is_solar_sector(sector):
                            solar_share_pct = solar_shares.get(sector, 0.0)
                            net_value = sector_value - (sector_value * solar_share_pct / 100.0)
                            row[sector] = net_value
                            net_total += net_value
                        else:
                            # For solar sectors, keep the absolute value
                            row[sector] = abs(sector_value)
                            net_total += abs(sector_value)

                row['Total'] = net_total

            elif demand_type == "onGrid":
                # ON GRID DEMAND: Net demand + T&D losses
                net_total = 0

                for sector in sectors:
                    if row[sector] is not None:
                        sector_value = float(row[sector])

                        # For non-solar sectors, subtract their solar share
                        if not is_solar_sector(sector):
                            solar_share_pct = solar_shares.get(sector, 0.0)
                            net_value = sector_value - (sector_value * solar_share_pct / 100.0)
                            row[sector] = net_value
                            net_total += net_value
                        else:
                            # For solar sectors, keep the absolute value
                            row[sector] = abs(sector_value)
                            net_total += abs(sector_value)

                # Apply T&D losses to net demand
                td_percentage = calculate_td_loss_percentage(year, td_loss_points)
                td_losses = net_total * (td_percentage / (1 - td_percentage))
                final_total = net_total + td_losses

                row['Net Total'] = net_total
                row['T&D Loss (%)'] = td_percentage
                row['T&D Losses'] = td_losses
                row['Total'] = final_total

        return {"success": True, "data": consolidated_data}

    except HTTPException:
        raise
    except Exception as error:
        logger.error(f"Error generating consolidated data: {error}")
        raise HTTPException(status_code=500, detail="Failed to fetch consolidated data.")


@router.post("/save-consolidated")
async def save_consolidated(request: SaveConsolidatedRequest):
    """
    Save consolidated results to Excel file.

    Args:
        request: Consolidated data to save

    Returns:
        dict: Success message with file path
    """
    if not request.projectPath or not request.scenarioName or not request.data:
        raise HTTPException(status_code=400, detail="Invalid or missing data for saving.")

    try:
        scenario_results_path = (
            Path(request.projectPath) / "results" / "demand_forecasts" / request.scenarioName
        )
        scenario_results_path.mkdir(parents=True, exist_ok=True)

        output_file_path = scenario_results_path / "Consolidated_Results.xlsx"

        # Process data: Convert T&D Loss (%) to percentage string
        processed_data = []
        for row in request.data:
            new_row = dict(row)
            if 'T&D Loss (%)' in new_row and new_row['T&D Loss (%)'] is not None:
                new_row['T&D Loss (%)'] = f"{float(new_row['T&D Loss (%)']) * 100:.2f}%"
            processed_data.append(new_row)

        # Create workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Consolidated Data"

        # Write headers
        headers = list(processed_data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Write data
        for row_idx, row_data in enumerate(processed_data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        # Save workbook
        workbook.save(output_file_path)
        workbook.close()

        logger.info(f"✅ Consolidated results saved to: {output_file_path}")

        return {
            "success": True,
            "message": "File saved successfully!",
            "path": str(output_file_path)
        }

    except Exception as error:
        logger.error(f"❌ FATAL ERROR during file save operation: {error}")
        raise HTTPException(status_code=500, detail="Failed to save the Excel file.")
